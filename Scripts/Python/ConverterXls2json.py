from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
import json

def NodeLayerProcessing(List):
    print("Enter Function")
    ExitList=[]
    print(List,type(List))
    for layer in List.split(","):
        layer=layer.replace(" ","")
        ExitList.append(layer)
    # for i in List:
    #     print(i)
    print("End function")
    return ExitList

def DirectLayerProcessing(Lists):
    # print("Enter Function")
    ExitList=[]
    QueryDict={}
    # print(Lists)
    for i in Lists:
        if i != None:
            # print(i,type(i))
            for j in i.split(","):
                j=j.replace(" ", "")
                print("Name:",j)
                QueryName=j.split("_") # budapest-bus_CL -> list [budapest-bus,CL]
                print("QueryName:",QueryName)
                print("QueryDict.keys()")
                print(QueryDict.keys())
                print("QueryName[1]")
                print(QueryName[1])
                if QueryName[1] in QueryDict.keys():
                    QueryDict[QueryName[1]].append(j)
                    # print("Key:",QueryName[1],"\tList:",QueryDict[QueryName[1]])
                else:
                    QueryDict[QueryName[1]]=[]
                    QueryDict[QueryName[1]].append(j)
                    # print("Key:",QueryName[1],"\tList:",QueryDict[QueryName[1]])
    for query in QueryDict.keys():
        # print(query,QueryDict[query])
        ExitList.append(QueryDict[query])
    # print("End function")
    # ExitList=[
    #                 ["montreal-bus_CDtest","montreal-metro_CDtest"],
    #                 ["montreal-bus_CLtest","montreal-metro_CLtest"]
    return (ExitList)

def WriteToFile(Data:dict,ExitPath:str) -> None:
    f = open(ExitPath, "w",encoding='utf-8')
    var=json.dumps(Data, ensure_ascii=False)
    # print("var",var)
    # print("var",type(var))
    f.write(var)
    f.close()

def ReadFile(path:str) -> dict:
    LetterHeaders={   
        "City":"A",
        "name":"B",
        "DirectStyleURL":"C",
        "NodeStyleURL":"D",
        "Coords-Lat":"E",
        "Coords-Lon":"F",
        "Zoom":"G",
        "NumTransitSystems":"H",
        "NodeLayers":"I",
        "Bus-DirectLayers":"J",
        "Bus-NumStops":"K",
        "Bus-NumLines":"L",
        "Bus-AvgDisStops":"M",
        "Train-DirectLayers":"N",
        "Train-NumStops":"O",
        "Train-NumLines":"P",
        "Train-AvgDisStops":"Q",
        "Metro-DirectLayers":"R",
        "Metro-NumStops":"S",
        "Metro-NumLines":"T",
        "Metro-AvgDisStops":"U",
        "Tram-DirectLayers":"V",
        "Tram-NumStops":"W",
        "Tram-NumLines":"X",
        "Tram-AvgDisStops":"Y",
        "Others-DirectLayers":"Z",
        "Others-NumStops":"AA",
        "Others-NumLines":"AB",
        "Others-AvgDisStops":"AC",
        "NumBoroughs":"AD",
        "AreaSqKm":"AE",
        "PopulationMillion":"AF",
        "DensityPersonSqKm":"AG",
        "YearOfStats":"AH",
        "SourceGTFS":"AI",
        "DateUpdatedGTFS":"AJ",
        "GridLayers":"AK",
        "GridStyleURL":"AL",
        "Flag":"AM",
	}

    ExitDict={"City": {}}

    wb = load_workbook(path)
    # print(wb.sheetnames)
    WorkSheet=wb.active
    # print(WorkSheet['A1'].value)
    # DATA BEGINS IN LINE 3
    CellCondition=True
    RowCount=3
    while(CellCondition):
        ExcelCoord="A"+str(RowCount)
        # print(ExcelCoord,WorkSheet[ExcelCoord].value)
        if WorkSheet[ExcelCoord].value == None:
            CellCondition=False
        else:
            RowCount+=1
   
    for Row in range(3,RowCount):
        ExcelCityCoord="A"+str(Row)
        CityEnglishName=WorkSheet[ExcelCityCoord].value
        TransitionDictionary={}
        for key in LetterHeaders.keys():
            letter=LetterHeaders[key]
            ExcelCoord=letter+str(Row)
            TransitionDictionary[key]=WorkSheet[ExcelCoord].value
            # print(key,"\t=",WorkSheet[ExcelCoord].value)

        DirectList=DirectLayerProcessing([TransitionDictionary["Bus-DirectLayers"],TransitionDictionary["Train-DirectLayers"],TransitionDictionary["Metro-DirectLayers"],TransitionDictionary["Tram-DirectLayers"],TransitionDictionary["Others-DirectLayers"],TransitionDictionary["GridLayers"]])
        NodeList=NodeLayerProcessing(TransitionDictionary["NodeLayers"]+","+TransitionDictionary["GridLayers"])

        # ListNode=[]
        # for i in TransitionDictionary["NodeLayers"].split(","):
        #    ListNode.append(i)
        
        # print("listTest",listTest,type(listTest))
        ExitDict["City"][CityEnglishName]={
                "name": TransitionDictionary["name"],
                "flag": TransitionDictionary["Flag"],
                "AreaSqKm": TransitionDictionary["AreaSqKm"],
                "PopulationMillion": TransitionDictionary["PopulationMillion"],
                "DensityPersonSqKm": TransitionDictionary["DensityPersonSqKm"],
                "NumBoroughs": TransitionDictionary["NumBoroughs"],
                "NumTransitSystems": TransitionDictionary["NumTransitSystems"],
                "YearOfStats":TransitionDictionary["YearOfStats"],
                "SourceGTFS":TransitionDictionary["SourceGTFS"],
                "DateUpdatedGTFS":TransitionDictionary["DateUpdatedGTFS"],
                "TransitSystems":[
                    {
                        "Type": "Bus",
                        "NumStops": TransitionDictionary["Bus-NumStops"],
                        "NumLines": TransitionDictionary["Bus-NumLines"],
                        "AvgDisStops": TransitionDictionary["Bus-AvgDisStops"]
                    },
                    {
                        "Type": "Train",
                        "NumStops": TransitionDictionary["Train-NumStops"],
                        "NumLines": TransitionDictionary["Train-NumLines"],
                        "AvgDisStops": TransitionDictionary["Train-AvgDisStops"]
                    },
                    {
                        "Type": "Metro",
                        "NumStops": TransitionDictionary["Metro-NumStops"],
                        "NumLines": TransitionDictionary["Metro-NumLines"],
                        "AvgDisStops": TransitionDictionary["Metro-AvgDisStops"]
                    },
                    {
                        "Type": "Tram",
                        "NumStops": TransitionDictionary["Tram-NumStops"],
                        "NumLines": TransitionDictionary["Tram-NumLines"],
                        "AvgDisStops": TransitionDictionary["Tram-AvgDisStops"]
                    },
                    {
                        "Type": "Others",
                        "NumStops": TransitionDictionary["Others-NumStops"],
                        "NumLines": TransitionDictionary["Others-NumLines"],
                        "AvgDisStops": TransitionDictionary["Others-AvgDisStops"]
                    }
                ],
                "DirectStyleURL": TransitionDictionary["DirectStyleURL"],
                "NodeStyleURL": TransitionDictionary["NodeStyleURL"],
                "DirectLayers": DirectList,
                "NodeLayers": NodeList,
                "GridStyleURL": TransitionDictionary["GridStyleURL"],
                "Coords": [
                    TransitionDictionary["Coords-Lat"],
                    TransitionDictionary["Coords-Lon"]
                ],
                "Zoom": TransitionDictionary["Zoom"]
            }
        

    return ExitDict														


if __name__ =="__main__":
    PathToTheFile="Data/DatabaseCitys.xlsx"
    Data=ReadFile(path=PathToTheFile)
    # print("\n"*10)
    # print(Data)
    WriteToFile(Data=Data,ExitPath="Data/CityMetrics.json")
    # for CityJs in Data['City']:
    #     print("\n"*10)
    #     print("CityJs",CityJs)
    #     print(Data['City'][CityJs])