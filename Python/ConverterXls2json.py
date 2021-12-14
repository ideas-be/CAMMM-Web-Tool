from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
import json

def DirectLayerProcessing(Lists):
    print("Enter Function")
    print(Lists)
    for i in Lists:
        if i != None:
            print(i,type(i))
            for j in i.split(","):
                print(j)
    print("End function")

def writeToFile(Data,ExitPath):
    f = open(ExitPath, "w")
    f.write(json.dumps(Data))
    f.close()

def ReadFile(path):
    LetterHeaders={   
        "City":"A",
        "name":"B",
        "DirectStyleURL":"C",
        "NodeStyleURL":"D",
        "Coords-Lat":"E",
        "Coords-Lon":"F",
        "Zoom":"G",
        "NumTransitSystems":"H",
        "NodeLayers":"I",
        "Bus-DirectLayers":"J",
        "Bus-NumStops":"K",
        "Bus-NumLines":"L",
        "Bus-AvgDisStops":"M",
        "Train-DirectLayers":"N",
        "Train-NumStops":"O",
        "Train-NumLines":"P",
        "Train-AvgDisStops":"Q",
        "Metro-DirectLayers":"R",
        "Metro-NumStops":"S",
        "Metro-NumLines":"T",
        "Metro-AvgDisStops":"U",
        "Tram-DirectLayers":"V",
        "Tram-NumStops":"W",
        "Tram-NumLines":"X",
        "Tram-AvgDisStops":"Y",
        "Others-DirectLayers":"Z",
        "Others-NumStops":"AA",
        "Others-NumLines":"AB",
        "Others-AvgDisStops":"AC",
        "NumBoroughs":"AD",
        "AreaSqKm":"AE",
        "PopulationMillion":"AF",
        "DensityPersonSqKm":"AG"
    }

    ExitDict={"City": {}}

    wb = load_workbook(path)
    # print(wb.sheetnames)
    WorkSheet=wb.active
    # print(WorkSheet['A1'].value)
    # DATA BEGINS IN LINE 3
    CellCondition=True
    RowCount=3
    while(CellCondition):
        ExcelCoord="A"+str(RowCount)
        # print(ExcelCoord,WorkSheet[ExcelCoord].value)
        if WorkSheet[ExcelCoord].value == None:
            CellCondition=False
        else:
            RowCount+=1
   
    for Row in range(3,RowCount):
        ExcelCityCoord="A"+str(Row)
        CityEnglishName=WorkSheet[ExcelCityCoord].value
        TransitionDictionary={}
        for key in LetterHeaders.keys():
            letter=LetterHeaders[key]
            ExcelCoord=letter+str(Row)
            TransitionDictionary[key]=WorkSheet[ExcelCoord].value
            # print(key,"\t=",WorkSheet[ExcelCoord].value)

        DirectLayerProcessing([TransitionDictionary["Bus-DirectLayers"],TransitionDictionary["Train-DirectLayers"],TransitionDictionary["Metro-DirectLayers"],TransitionDictionary["Tram-DirectLayers"],TransitionDictionary["Others-DirectLayers"]])

        # ListNode=[]
        # for i in TransitionDictionary["NodeLayers"].split(","):
        #    ListNode.append(i)
        
        # print("listTest",listTest,type(listTest))
        ExitDict["City"][CityEnglishName]={
            TransitionDictionary["City"]: {
                "name": TransitionDictionary["name"],
                "AreaSqKm": TransitionDictionary["AreaSqKm"],
                "PopulationMillion": TransitionDictionary["PopulationMillion"],
                "DensityPersonSqKm": TransitionDictionary["DensityPersonSqKm"],
                "NumBoroughs": TransitionDictionary["NumBoroughs"],
                "NumTransitSystems": TransitionDictionary["NumTransitSystems"],
                "TransitSystems":[
                    {
                        "Type": "Bus",
                        "NumStops": TransitionDictionary["Bus-NumStops"],
                        "NumLines": TransitionDictionary["Bus-NumLines"],
                        "AvgDisStops": TransitionDictionary["Bus-AvgDisStops"]
                    },
                    {
                        "Type": "Train",
                        "NumStops": TransitionDictionary["Train-NumStops"],
                        "NumLines": TransitionDictionary["Train-NumLines"],
                        "AvgDisStops": TransitionDictionary["Train-AvgDisStops"]
                    },
                    {
                        "Type": "Metro",
                        "NumStops": TransitionDictionary["Metro-NumStops"],
                        "NumLines": TransitionDictionary["Metro-NumLines"],
                        "AvgDisStops": TransitionDictionary["Metro-AvgDisStops"]
                    },
                    {
                        "Type": "Tram",
                        "NumStops": TransitionDictionary["Tram-NumStops"],
                        "NumLines": TransitionDictionary["Tram-NumLines"],
                        "AvgDisStops": TransitionDictionary["Tram-AvgDisStops"]
                    },
                    {
                        "Type": "Others",
                        "NumStops": TransitionDictionary["Others-NumStops"],
                        "NumLines": TransitionDictionary["Others-NumLines"],
                        "AvgDisStops": TransitionDictionary["Others-AvgDisStops"]
                    }
                ],
                "DirectStyleURL": TransitionDictionary["DirectStyleURL"],
                "NodeStyleURL": TransitionDictionary["NodeStyleURL"],
                "DirectLayers": [
                    ["montreal-bus_CD","montreal-metro_CD"],
                    ["montreal-bus_CL","montreal-metro_CL"]
                ],
                "NodeLayers": [
                    "montreal_N_CD",
                    "montreal_N_CL"
                ],
                "Coords": [
                    TransitionDictionary["Coords-Lat"],
                    TransitionDictionary["Coords-Lon"]
                ],
                "Zoom": TransitionDictionary["Zoom"]
            }
        }

    return ExitDict														


if __name__ =="__main__":
    PathToTheFile="DatabaseCitys.xlsx"
    Data=ReadFile(path=PathToTheFile)
    print("\n"*10)
    print(Data)
    writeToFile(Data=Data,ExitPath="Sample.json")
    # for CityJs in Data['City']:
    #     print("\n"*10)
    #     print("CityJs",CityJs)
    #     print(Data['City'][CityJs])